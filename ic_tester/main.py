from turtle import width
import openpyxl 
import math
import tkinter as tk
from tkinter import messagebox
from utilites import filter_row, save_to_file

def process_readings() -> None:
    try:
        readings = list(map(float, map(str.strip, entry.get().split(','))))
        indexer = 0
        is_matching = True
        for key, value in ic_saved_dict.items():
            while indexer <= len(value) - 1:
                if readings[indexer] != value[indexer]:
                    is_matching = False
                    break
                indexer += 1
            if is_matching:
                messagebox.showinfo('Test Result', f'IC Name is {key}')
                break
            is_matching = True
            indexer = 0
        if not is_matching:
            messagebox.showinfo('Test result', "Your IC wasn't found")
    except Exception as err:
        messagebox.showerror('Input error', 'Please enput the readings in the right format')
        print(f'Error: {err}')

if __name__ == "__main__":
    root = tk.Tk()
    root.title("IC Tester")
    root.geometry("900x300")
    wb = openpyxl.load_workbook('./data/embedded_sysetms_ICS_readings.xlsx')
    ws = wb.active
    rows_count = ws.max_row
    ics_count = math.ceil(rows_count / 5)
    rows_values = list(ws.iter_rows(values_only=True))
    current_ic_index = 0
    ic_saved_dict = {}
    while current_ic_index < ics_count:
        indexer = 5 * current_ic_index
        names_row = rows_values[indexer]
        pins_row = rows_values[indexer + 1]
        readings_row = rows_values[indexer + 2]
        normalized_row = rows_values[indexer + 3]
        ic_saved_dict[names_row[0]] = filter_row(normalized_row[2:])
        current_ic_index += 1

    save_to_file(ic_saved_dict)
    label = tk.Label(root, text="Enter IC readings: ")
    label.pack(pady=10)
    entry = tk.Entry(root, width=80)
    entry.pack(pady=5)
    button = tk.Button(root, text="Submit", command=process_readings)
    button.pack(pady=10)
    root.mainloop()
